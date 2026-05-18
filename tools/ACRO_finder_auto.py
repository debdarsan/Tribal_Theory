import re
from langchain_community.vectorstores import Chroma
import chromadb
from chromadb import Settings
from langchain_openai import OpenAIEmbeddings
from dotenv import load_dotenv
import pandas as pd
import datetime
import os
import csv
from openpyxl import load_workbook

load_dotenv()

PERSIST_VECTORDB_DIR = "Chroma_VectorStore"
SECTIONS = ["Glossary", "Glossary of Terms", "Index"]

def find_acronyms_and_meanings(text):
    # Find potential acronyms (two or more consecutive uppercase letters)
    acronyms = re.findall(r'\b[A-Z]{2,}\b', text)

    results = {}
    for acronym in acronyms:
        # Create a regex pattern for each letter in the acronym, followed by any number of alphabets
        acronym_regex = r'\b' + r'\s+'.join([letter + r'[a-z]*' for letter in acronym]) + r'\b'
        match = re.search(acronym_regex, text, re.IGNORECASE)

        if match:
            # Extract the full form corresponding to the acronym
            full_form = match.group(0)
            results[acronym] = full_form

    return results

def find_and_write_acronyms_to_csv(markdown_filename, output_filename):
    with open(markdown_filename, 'r', encoding="utf-8") as file:
        content = file.read()
    acronyms_and_meanings_advanced = find_acronyms_and_meanings(content)
    print(acronyms_and_meanings_advanced)
    with open(output_filename, 'w', newline='', encoding='utf-8') as csvfile: 
        # Create a csv writer object 
        csvwriter = csv.writer(csvfile) 

        # Write the headers
        csvwriter.writerow(["Acronyms", "Meanings"])

        # Write the data to the CSV file
        for key, value in acronyms_and_meanings_advanced.items():
            csvwriter.writerow([key, value])

def load_vectorstore(PERSIST_VECTORDB_DIR):
    vdb = None
    record_count = 0
    # define an empty list of ids
    existing_ids = []
    CHROMA_SETTINGS = Settings(
                    persist_directory=PERSIST_VECTORDB_DIR,
                    # Prevent anonymous data collection by Chroma
                    anonymized_telemetry=False,
                )
    chroma_client = chromadb.PersistentClient(
        settings=CHROMA_SETTINGS, path=PERSIST_VECTORDB_DIR,   
    )
    embeddings = OpenAIEmbeddings()
    # define Chroma Vector DB
    vdb = Chroma(
        persist_directory=PERSIST_VECTORDB_DIR,
        embedding_function=embeddings,
        client_settings=CHROMA_SETTINGS,
        client=chroma_client,
    )
    try:
        record_count = vdb._collection.count()
        if record_count > 0:
            existing_ids = [str(id) for id in vdb._collection.get()["ids"]]
            metadata = [m for m in vdb._collection.get()["metadatas"]]
            record_count = vdb._collection.count()
        else:
            print(f"... New Chroma VDB is to be created at directory: {PERSIST_VECTORDB_DIR}",)
    except Exception as e:
        raise
    return vdb, existing_ids, metadata, record_count

def find_filenames_in_vectorstore_and_write(output_folder):
    vdb, existing_ids, metadata, record_count = load_vectorstore(PERSIST_VECTORDB_DIR)
    # Create an empty set to store the unique file names
    file_names = set()

    # Loop through the metadata
    for item in metadata:
        # Get the file name of the item
        file_name = item["File"]
        file_names.add(file_name)
        
    # Create a dataframe to store the file names
    df = pd.DataFrame({"File Name": sorted(list(file_names))})
    df.index = df.index + 1
    df.index.name = "No."
    # Generate the current date in the format "dd-mon-yy"
    current_date = datetime.datetime.now().strftime("%d-%b-%y")
    file_name = output_folder + os.sep + f"Docs in Chroma on {current_date}.xlsx"

    # Save the dataframe as an xlsx file with the modified file name
    df.to_excel(file_name, index=True)

    # Print the number and list of unique file names
    print(f"There are {len(file_names)} unique file names in the database.")
    print("The file names are:")
    for i, file_name in enumerate(file_names, 1):
        print(f"{i}. {file_name}")

    return file_names

def extract_markdown_sections(md_content, sections_to_be_removed):
    lines = md_content.splitlines()
    sections = []
    current_section = None
    current_text = []

    for line in lines:
        # Remove '#' and '*' from the line and strip spaces
        cleaned_line = line.replace('#', '').replace('*', '').strip()
        
        if any(cleaned_line == section for section in sections_to_be_removed):
            # Start a new section
            if current_section:
                sections.append({'section': current_section, 'section_text': '\n'.join(current_text)})
                current_text = []
            current_section = cleaned_line
        elif line.startswith(('#', '*')) and current_section:
            # End the current section
            sections.append({'section': current_section, 'section_text': '\n'.join(current_text)})
            current_section = None
            current_text = []
        else:
            # Add line to section text or final content
            if current_section:
                current_text.append(line)

    # Add the last section (if any)
    if current_section:
        sections.append({'section': current_section, 'section_text': '\n'.join(current_text)})

    return sections

def list_files_in_folder(folder_path):
    # Check if the provided path is a directory
    if not os.path.isdir(folder_path):
        print("Error: The provided path is not a directory.")
        return None
    
    # Get the list of filenames in the folder
    filenames = os.listdir(folder_path)
    
    # Return the list of filenames
    return filenames

def extract_filenames_from_excel(folder_path):
    names = []

    # Iterate through files in the folder
    for filename in os.listdir(folder_path):
        if filename.startswith('Docs in Chroma on') and filename.endswith('.xlsx'):
            filepath = os.path.join(folder_path, filename)

            # Load the Excel file
            try:
                workbook = load_workbook(filepath)
                sheet = workbook.active

                # Assuming the "name" column is in the second column (B)
                for row in sheet.iter_rows(values_only=True):
                    name = row[1]  # Index 1 corresponds to the "name" column
                    if name:
                        names.append(name)
                
                workbook.close()
            except Exception as e:
                print(f"Error processing file '{filename}': {e}")

    return names

old_filenames = extract_filenames_from_excel("./DocumentSource")
new_filenames = find_filenames_in_vectorstore_and_write("./DocumentSource")
files_in_ingestion = list_files_in_folder("./ingestion")
print(old_filenames, new_filenames)